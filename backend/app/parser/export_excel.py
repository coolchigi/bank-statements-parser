#!/usr/bin/env python3
"""
export_excel.py
---------------
Writes a list of Transaction objects to a formatted Excel workbook.

Sheet 1 — "Transactions": every transaction, one row each
Sheet 2 — "Summary by Category": totals grouped by category
"""

from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from app.parser.parse_statements import Transaction

# ─── STYLES ───────────────────────────────────────────────────────────────────

HEADER_FILL     = PatternFill("solid", fgColor="1F3864")
HEADER_FONT     = Font(bold=True, color="FFFFFF", size=11)
ROW_FILL_ODD    = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN   = PatternFill("solid", fgColor="DCE6F1")
WITHDRAWAL_FONT = Font(color="C00000")           # dark red
DEPOSIT_FONT    = Font(color="375623")           # dark green
GRAND_FILL      = PatternFill("solid", fgColor="1F3864")
GRAND_FONT      = Font(bold=True, color="FFFFFF")

CURRENCY_FMT = '$#,##0.00'
DATE_FMT     = 'DD-MMM-YYYY'

# ─── TRANSACTIONS SHEET ───────────────────────────────────────────────────────

def write_transactions_sheet(wb: openpyxl.Workbook, transactions: List[Transaction]):
    ws = wb.active
    ws.title = "Transactions"

    headers    = ["Date", "Type", "Amount", "Category",
                  "Description", "Merchant", "Statement Period", "Balance"]
    col_widths = [14,      14,     14,        24,
                  58,           32,          26,                  14]

    # ── Header row ──
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Data rows ──
    for row_idx, t in enumerate(transactions, start=2):
        fill        = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        amount_font = WITHDRAWAL_FONT if t.direction == "Withdrawal" else DEPOSIT_FONT

        def c(col, value, fmt=None, font=None, align="left"):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            if fmt:
                cell.number_format = fmt
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal=align)
            return cell

        c(1, t.date,                                  fmt=DATE_FMT,     align="center")
        c(2, t.direction,                                                align="center")
        c(3, float(t.amount),             fmt=CURRENCY_FMT, font=amount_font, align="right")
        c(4, t.category)
        c(5, t.description)
        c(6, t.merchant)
        c(7, t.statement_period,                                         align="center")
        c(8, float(t.balance) if t.balance is not None else None,
              fmt=CURRENCY_FMT if t.balance is not None else None,       align="right")


# ─── SUMMARY SHEET ────────────────────────────────────────────────────────────

def write_summary_sheet(wb: openpyxl.Workbook, transactions: List[Transaction]):
    ws = wb.create_sheet("Summary by Category")

    headers    = ["Category", "Total Withdrawals", "Total Deposits", "Net", "# Transactions"]
    col_widths = [26,          22,                  22,               22,    18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Aggregate ──
    summary = defaultdict(lambda: {"withdrawals": Decimal(0), "deposits": Decimal(0), "count": 0})
    for t in transactions:
        entry = summary[t.category]
        entry["count"] += 1
        if t.direction == "Withdrawal":
            entry["withdrawals"] += t.amount
        else:
            entry["deposits"] += t.amount

    sorted_cats = sorted(summary.items(), key=lambda x: x[1]["withdrawals"], reverse=True)

    for row_idx, (cat, data) in enumerate(sorted_cats, start=2):
        fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        net  = data["deposits"] - data["withdrawals"]

        ws.cell(row=row_idx, column=1, value=cat).fill = fill

        for col_idx, (val, font) in enumerate([
            (float(data["withdrawals"]), WITHDRAWAL_FONT),
            (float(data["deposits"]),   DEPOSIT_FONT),
            (float(net),                DEPOSIT_FONT if net >= 0 else WITHDRAWAL_FONT),
        ], start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = CURRENCY_FMT
            cell.font  = font
            cell.fill  = fill
            cell.alignment = Alignment(horizontal="right")

        cnt = ws.cell(row=row_idx, column=5, value=data["count"])
        cnt.fill      = fill
        cnt.alignment = Alignment(horizontal="center")

    # ── Grand total row ──
    total_row = len(sorted_cats) + 2
    total_w   = sum(d["withdrawals"] for d in summary.values())
    total_d   = sum(d["deposits"]    for d in summary.values())
    total_net = total_d - total_w
    total_cnt = sum(d["count"]       for d in summary.values())

    for col_idx, val in enumerate(
        ["TOTAL", float(total_w), float(total_d), float(total_net), total_cnt], start=1
    ):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.fill = GRAND_FILL
        cell.font = GRAND_FONT
        if col_idx in (2, 3, 4):
            cell.number_format = CURRENCY_FMT
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 5) else "right")


# ─── MAIN EXPORT ENTRY POINT ──────────────────────────────────────────────────

import io as _io

def export(transactions: List[Transaction], output_path=None) -> bytes:
    """
    Write transactions to an Excel workbook.

    Args:
        transactions: list of Transaction objects
        output_path:  optional Path to save to disk; if None, returns raw bytes only

    Returns:
        Raw Excel bytes (always)
    """
    wb = openpyxl.Workbook()
    write_transactions_sheet(wb, transactions)
    write_summary_sheet(wb, transactions)

    buf = _io.BytesIO()
    wb.save(buf)
    raw_bytes = buf.getvalue()

    if output_path is not None:
        with open(output_path, "wb") as f:
            f.write(raw_bytes)

    return raw_bytes
